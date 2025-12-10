from datetime import datetime, date
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from machines.models import Machine
from openpyxl import load_workbook
from references.models import ReferenceItem

EXCEL_COLUMNS = {
    "serial_number": "Зав. № машины",
    "machine_model": "Модель техники",
    "engine_model": "Модель двигателя",
    "engine_serial_number": "Зав. № двигателя",
    "transmission_model": "Модель трансмиссии",
    "transmission_serial_number": "Зав. № трансмиссии",
    "drive_axle_model": "Модель ведущего моста",
    "drive_axle_serial_number": "Зав. № ведущего моста",
    "steer_axle_model": "Модель управляемого моста",
    "steer_axle_serial_number": "Зав. № управляемого моста",
    "contract_number_and_date": "Договор поставки №, дата",
    "shipment_date": "Дата отгрузки с завода",
    "consignee": "Грузополучатель (конечный потребитель)",
    "delivery_address": "Адрес поставки (эксплуатации)",
    "options": "Комплектация (доп. опции)",
}


class Command(BaseCommand):
    help = "Импорт машин из Excel (data.xlsx) в базу данных"

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            help="Путь к XLSX-файлу (по умолчанию BASE_DIR / 'data.xlsx')",
        )
        parser.add_argument(
            "--update",
            action="store_true",
            help="Обновлять существующие записи по serial_number (update_or_create)",
        )

    def handle(self, *args, **options):
        path_arg = options.get("path")
        update = options.get("update")

        if path_arg:
            xlsx_path = Path(path_arg)
        else:
            xlsx_path = Path(settings.BASE_DIR) / "data.xlsx"

        if not xlsx_path.exists():
            raise CommandError(f"Файл не найден: {xlsx_path}")

        self.stdout.write(f"Читаю файл: {xlsx_path}")

        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.stdout.write(self.style.WARNING("Файл пустой"))
            return

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        header_index = {name: idx for idx, name in enumerate(headers)}

        missing_headers = [
            excel_name
            for excel_name in EXCEL_COLUMNS.values()
            if excel_name not in header_index
        ]
        if missing_headers:
            self.stdout.write(
                self.style.WARNING(
                    "ВНИМАНИЕ: в файле не найдены некоторые ожидаемые колонки:\n"
                    + "\n".join(f"  - {h}" for h in missing_headers)
                )
            )

        def get_cell(row, logical_name):
            excel_col_name = EXCEL_COLUMNS.get(logical_name)
            if not excel_col_name:
                return None
            idx = header_index.get(excel_col_name)
            if idx is None:
                return None
            return row[idx]

        def parse_date(value):
            if value is None:
                return None

            if isinstance(value, date):
                return value
            if isinstance(value, datetime):
                return value.date()

            s = str(value).strip()
            if not s:
                return None

            for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue

            self.stdout.write(
                self.style.WARNING(f"Не удалось распознать дату '{s}', пропускаю.")
            )
            return None

        def get_or_create_ref(category, raw_value, field_verbose, serial_number):
            if raw_value is None:
                raise CommandError(
                    f"{field_verbose} пустое для машины с серийным номером '{serial_number}'"
                )

            name = str(raw_value).strip()
            if not name:
                raise CommandError(
                    f"{field_verbose} пустое для машины с серийным номером '{serial_number}'"
                )

            ref, _ = ReferenceItem.objects.get_or_create(
                category=category,
                name=name,
                defaults={"description": ""},
            )
            return ref

        created = 0
        updated_count = 0
        skipped = 0
        row_num = 1

        with transaction.atomic():
            for row in rows[1:]:
                row_num += 1
                if not any(row):
                    continue

                serial_raw = get_cell(row, "serial_number")
                if serial_raw is None:
                    skipped += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"Строка {row_num}: нет значения в колонке "
                            f"'{EXCEL_COLUMNS['serial_number']}', пропускаю."
                        )
                    )
                    continue

                serial_number = str(serial_raw).strip()
                if not serial_number:
                    skipped += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"Строка {row_num}: пустой серийный номер, пропускаю."
                        )
                    )
                    continue

                machine_model_raw = get_cell(row, "machine_model")
                engine_model_raw = get_cell(row, "engine_model")
                transmission_model_raw = get_cell(row, "transmission_model")
                drive_axle_model_raw = get_cell(row, "drive_axle_model")
                steer_axle_model_raw = get_cell(row, "steer_axle_model")

                machine_model = get_or_create_ref(
                    ReferenceItem.Category.MACHINE_MODEL,
                    machine_model_raw,
                    "Модель техники",
                    serial_number,
                )
                engine_model = get_or_create_ref(
                    ReferenceItem.Category.ENGINE_MODEL,
                    engine_model_raw,
                    "Модель двигателя",
                    serial_number,
                )
                transmission_model = get_or_create_ref(
                    ReferenceItem.Category.TRANSMISSION_MODEL,
                    transmission_model_raw,
                    "Модель трансмиссии",
                    serial_number,
                )
                drive_axle_model = get_or_create_ref(
                    ReferenceItem.Category.DRIVE_AXLE_MODEL,
                    drive_axle_model_raw,
                    "Модель ведущего моста",
                    serial_number,
                )
                steer_axle_model = get_or_create_ref(
                    ReferenceItem.Category.STEER_AXLE_MODEL,
                    steer_axle_model_raw,
                    "Модель управляемого моста",
                    serial_number,
                )

                defaults = {
                    "machine_model": machine_model,
                    "engine_model": engine_model,
                    "engine_serial_number": str(
                        get_cell(row, "engine_serial_number") or ""
                    ).strip(),
                    "transmission_model": transmission_model,
                    "transmission_serial_number": str(
                        get_cell(row, "transmission_serial_number") or ""
                    ).strip(),
                    "drive_axle_model": drive_axle_model,
                    "drive_axle_serial_number": str(
                        get_cell(row, "drive_axle_serial_number") or ""
                    ).strip(),
                    "steer_axle_model": steer_axle_model,
                    "steer_axle_serial_number": str(
                        get_cell(row, "steer_axle_serial_number") or ""
                    ).strip(),
                    "contract_number_and_date": str(
                        get_cell(row, "contract_number_and_date") or ""
                    ).strip(),
                    "shipment_date": parse_date(get_cell(row, "shipment_date")),
                    "consignee": str(get_cell(row, "consignee") or "").strip(),
                    "delivery_address": str(
                        get_cell(row, "delivery_address") or ""
                    ).strip(),
                    "options": str(get_cell(row, "options") or "").strip(),
                }

                if update:
                    obj, created_flag = Machine.objects.update_or_create(
                        serial_number=serial_number,
                        defaults=defaults,
                    )
                    if created_flag:
                        created += 1
                    else:
                        updated_count += 1
                else:
                    if Machine.objects.filter(serial_number=serial_number).exists():
                        skipped += 1
                        self.stdout.write(
                            self.style.WARNING(
                                f"Строка {row_num}: машина с серийным номером "
                                f"'{serial_number}' уже есть, пропускаю "
                                f"(запусти с --update, чтобы обновлять)."
                            )
                        )
                        continue
                    Machine.objects.create(serial_number=serial_number, **defaults)
                    created += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Импорт завершён. Создано: {created}, обновлено: {updated_count}, "
                f"пропущено: {skipped}"
            )
        )
